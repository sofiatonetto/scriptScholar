"""
Coleta do Google Acadêmico — "geotecnologias" e "ensino médio integrado"
Saída: Excel (.xlsx) com células amarelas para artigos sem PDF
Instalar: pip install requests beautifulsoup4 tqdm openpyxl
"""

import time, random, csv, json, re, logging, argparse
from pathlib import Path
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from tqdm import tqdm

# ──────────────────────────────────────────────────────────────
# CONFIGURAÇÕES
# ──────────────────────────────────────────────────────────────
OUTPUT_DIR  = Path("resultados_geo_emi")
PDF_DIR     = OUTPUT_DIR / "pdfs"
XLSX_FILE   = OUTPUT_DIR / "metadados.xlsx"
JSON_FILE   = OUTPUT_DIR / "metadados.json"
PROGRESSO   = OUTPUT_DIR / "progresso.json"
LOG_FILE    = OUTPUT_DIR / "scraper.log"

DELAY_PAG_MIN    = 20
DELAY_PAG_MAX    = 40
DELAY_LOTE_MIN   = 6
DELAY_LOTE_MAX   = 12
PAGINAS_POR_LOTE = 5
MAX_VAZIAS       = 5

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
]

CAMPOS = ["titulo", "autores", "fonte", "ano", "citacoes",
          "resumo", "link", "pdf_url", "pdf_local", "acesso_manual"]

# ──────────────────────────────────────────────────────────────
# SETUP
# ──────────────────────────────────────────────────────────────
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
PDF_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────
# PROGRESSO
# ──────────────────────────────────────────────────────────────
def carregar_progresso() -> dict:
    if PROGRESSO.exists():
        with open(PROGRESSO, encoding="utf-8") as f:
            return json.load(f)
    return {"ultima_pagina": 0, "total": 0, "ids_vistos": []}

def salvar_progresso(pagina, total, ids_vistos):
    with open(PROGRESSO, "w", encoding="utf-8") as f:
        json.dump({
            "ultima_pagina": pagina,
            "total": total,
            "ids_vistos": ids_vistos,
            "atualizado": datetime.now().isoformat()
        }, f, indent=2)


# ──────────────────────────────────────────────────────────────
# SESSÃO HTTP
# ──────────────────────────────────────────────────────────────
def criar_sessao(proxy=None):
    s = requests.Session()
    s.headers.update({
        "User-Agent": random.choice(USER_AGENTS),
        "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "DNT": "1",
    })
    if proxy:
        s.proxies = {"http": proxy, "https": proxy}
    return s

def rotacionar_agente(sessao):
    sessao.headers["User-Agent"] = random.choice(USER_AGENTS)


# ──────────────────────────────────────────────────────────────
# EXTRAÇÃO
# ──────────────────────────────────────────────────────────────
def extrair_metadados(item) -> dict | None:
    try:
        titulo_tag = item.find("h3", class_="gs_rt")
        if not titulo_tag:
            return None
        titulo = titulo_tag.get_text(separator=" ", strip=True)
        titulo = re.sub(r'\[(PDF|BOOK|CITATION|HTML)\]', '', titulo).strip()
        if not titulo:
            return None

        link_artigo = ""
        link_a = titulo_tag.find("a")
        if link_a:
            link_artigo = link_a.get("href", "")

        meta_tag = item.find("div", class_="gs_a")
        autores, fonte, ano = "", "", ""
        if meta_tag:
            meta_txt = meta_tag.get_text(separator="|||", strip=True)
            partes = meta_txt.split("|||")
            autores = partes[0].strip() if partes else ""
            fonte   = partes[1].strip() if len(partes) > 1 else ""
            resto   = " ".join(partes)
            m = re.search(r'\b(20\d{2}|19\d{2})\b', resto)
            ano = m.group(0) if m else ""

        resumo_tag = item.find("div", class_="gs_rs")
        resumo = resumo_tag.get_text(strip=True) if resumo_tag else ""

        pdf_url = ""
        pdf_div = item.find("div", class_="gs_or_ggsm")
        if pdf_div:
            pdf_a = pdf_div.find("a")
            if pdf_a:
                pdf_url = pdf_a.get("href", "")

        citacoes = 0
        for a in item.find_all("a"):
            txt = a.get_text(strip=True)
            if txt.startswith(("Citado por", "Cited by")):
                m = re.search(r'\d+', txt)
                citacoes = int(m.group(0)) if m else 0
                break

        return {
            "titulo":        titulo,
            "autores":       autores,
            "fonte":         fonte,
            "ano":           ano,
            "citacoes":      citacoes,
            "resumo":        resumo,
            "link":          link_artigo,
            "pdf_url":       pdf_url,
            "pdf_local":     "",
            "acesso_manual": "",
        }
    except Exception as e:
        log.warning(f"Erro ao extrair: {e}")
        return None


# ──────────────────────────────────────────────────────────────
# BUSCA POR PÁGINA
# ──────────────────────────────────────────────────────────────
def buscar_pagina(sessao, query, start) -> tuple[list, str]:
    params = {
        "q":      query,
        "hl":     "pt",
        "start":  start,
        "as_sdt": "0,5",
        "as_yhi": 2025,
    }
    try:
        resp = sessao.get("https://scholar.google.com/scholar",
                          params=params, timeout=25)
    except requests.exceptions.RequestException as e:
        log.error(f"Erro de conexão: {e}")
        return [], "erro"

    if resp.status_code == 429:
        log.warning("⛔ 429 — bloqueio. Pausando 30 min...")
        time.sleep(30 * 60)
        return [], "bloqueio"

    if resp.status_code != 200:
        log.error(f"Status: {resp.status_code}")
        return [], "erro"

    soup = BeautifulSoup(resp.text, "html.parser")

    if ("please show you're not a robot" in resp.text.lower()
            or soup.find("form", {"id": "gs_captcha_f"})
            or "g-recaptcha" in resp.text):
        log.error(
            "\n" + "="*60 +
            "\n🛑 CAPTCHA!\n"
            "   1. Abra scholar.google.com no navegador\n"
            "   2. Resolva o CAPTCHA\n"
            "   3. Aguarde 2 horas\n"
            "   4. Rode o script de novo — retoma de onde parou\n" +
            "="*60
        )
        return [], "captcha"

    resultados = soup.find_all("div", class_="gs_r gs_or gs_scl")
    if not resultados:
        return [], "vazia"

    artigos = [a for item in resultados if (a := extrair_metadados(item))]
    return artigos, "ok"


# ──────────────────────────────────────────────────────────────
# BUSCA PRINCIPAL
# ──────────────────────────────────────────────────────────────
def buscar_tudo(sessao, query, meta) -> list:
    progresso   = carregar_progresso()
    inicio_pag  = progresso["ultima_pagina"]
    ids_vistos  = set(progresso.get("ids_vistos", []))
    todos       = []

    if PROGRESSO.exists() and inicio_pag > 0:
        if JSON_FILE.exists():
            with open(JSON_FILE, encoding="utf-8") as f:
                todos = json.load(f)
        log.info(f"▶ Retomando da pág {inicio_pag + 1} ({len(todos)} já coletados)")

    num_paginas  = max((meta // 10) + 5, 15)
    paginas_lote = 0
    vazias       = 0

    for pagina in range(inicio_pag, num_paginas):
        if len(todos) >= meta:
            log.info(f"✅ Meta de {meta} atingida!")
            break

        start = pagina * 10

        if paginas_lote > 0 and paginas_lote % PAGINAS_POR_LOTE == 0:
            minutos = random.uniform(DELAY_LOTE_MIN, DELAY_LOTE_MAX)
            log.info(f"\n⏸  Pausa: {minutos:.1f} min...\n")
            time.sleep(minutos * 60)
            rotacionar_agente(sessao)

        log.info(f"[Pág {pagina+1}/{num_paginas}] coletados={len(todos)}/{meta} | vazias={vazias}")

        artigos, status = buscar_pagina(sessao, query, start)

        if status == "captcha":
            break
        if status in ("bloqueio", "erro"):
            time.sleep(15)
            continue
        if status == "vazia":
            vazias += 1
            if vazias >= MAX_VAZIAS:
                log.info("Fim dos resultados.")
                break
            salvar_progresso(pagina + 1, len(todos), list(ids_vistos))
            paginas_lote += 1
            time.sleep(random.uniform(DELAY_PAG_MIN, DELAY_PAG_MAX))
            continue

        vazias = 0
        novos = []
        for a in artigos:
            if len(todos) >= meta:
                break
            chave = a["titulo"].lower()[:60]
            if chave not in ids_vistos:
                todos.append(a)
                novos.append(a)
                ids_vistos.add(chave)

        if novos:
            salvar_progresso(pagina + 1, len(todos), list(ids_vistos))
            # Salva JSON intermediário para retomada
            with open(JSON_FILE, "w", encoding="utf-8") as f:
                json.dump(todos, f, ensure_ascii=False, indent=2)
            log.info(f"  ✔ +{len(novos)} | Total: {len(todos)}")
        else:
            log.info(f"  Nenhum novo (duplicatas)")

        paginas_lote += 1
        espera = random.uniform(DELAY_PAG_MIN, DELAY_PAG_MAX)
        log.info(f"  Próxima em {espera:.0f}s...")
        time.sleep(espera)

    return todos


# ──────────────────────────────────────────────────────────────
# DOWNLOAD DE PDFs
# ──────────────────────────────────────────────────────────────
def nome_seguro(titulo, idx):
    nome = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", titulo)[:80].strip()
    return f"{idx:04d}_{nome}.pdf"

def baixar_pdf(sessao, url, destino):
    try:
        resp = sessao.get(url, timeout=30, stream=True, allow_redirects=True)
        ct = resp.headers.get("Content-Type", "")
        if resp.status_code != 200:
            return False
        if "pdf" not in ct and not url.lower().endswith(".pdf"):
            return False
        with open(destino, "wb") as f:
            for chunk in resp.iter_content(8192):
                f.write(chunk)
        return True
    except Exception:
        return False

def baixar_todos_pdfs(sessao, artigos):
    com_pdf = sum(1 for a in artigos if a.get("pdf_url"))
    log.info(f"Baixando PDFs ({com_pdf} disponíveis de {len(artigos)} artigos)...")
    salvos = 0
    for idx, artigo in enumerate(tqdm(artigos, desc="PDFs"), 1):
        if not artigo.get("pdf_url"):
            # Sem PDF disponível — marca link para acesso manual
            artigo["acesso_manual"] = artigo.get("link", "")
            continue
        destino = PDF_DIR / nome_seguro(artigo["titulo"], idx)
        if destino.exists():
            artigo["pdf_local"] = str(destino.resolve())
            salvos += 1
            continue
        if baixar_pdf(sessao, artigo["pdf_url"], destino):
            artigo["pdf_local"] = str(destino.resolve())
            artigo["acesso_manual"] = ""
            salvos += 1
            log.info(f"  ✔ {destino.name}")
        else:
            # Tinha link mas falhou — marca para acesso manual
            artigo["acesso_manual"] = artigo.get("link", "")
        time.sleep(random.uniform(2, 6))
    log.info(f"PDFs salvos: {salvos} | Acesso manual necessário: {len(artigos) - salvos}")
    return artigos


# ──────────────────────────────────────────────────────────────
# SALVAR XLSX COM CORES
# ──────────────────────────────────────────────────────────────
def salvar_xlsx(artigos):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl não instalado: pip install openpyxl")
        return

    ordenados = sorted(artigos, key=lambda a: int(a.get("ano") or 0), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Artigos"

    # Cabeçalho
    cabecalhos = ["#", "Título", "Autores", "Fonte", "Ano", "Citações",
                  "Resumo", "Link", "PDF URL", "PDF Local", "Acesso Manual"]
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, cab in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=cab)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30

    # Cores
    amarelo  = PatternFill("solid", fgColor="FFF2CC")  # sem PDF — acesso manual
    verde    = PatternFill("solid", fgColor="E2EFDA")  # PDF baixado
    branco   = PatternFill("solid", fgColor="FFFFFF")

    for i, a in enumerate(ordenados, 2):
        tem_pdf     = bool(a.get("pdf_local"))
        precisa_manual = bool(a.get("acesso_manual"))

        fill = verde if tem_pdf else (amarelo if precisa_manual else branco)

        valores = [
            i - 1,
            a.get("titulo", ""),
            a.get("autores", ""),
            a.get("fonte", ""),
            a.get("ano", ""),
            a.get("citacoes", 0),
            a.get("resumo", ""),
            a.get("link", ""),
            a.get("pdf_url", ""),
            a.get("pdf_local", ""),
            a.get("acesso_manual", ""),
        ]

        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col in [2, 3, 7]))

    # Larguras das colunas
    larguras = [5, 50, 30, 25, 8, 10, 60, 40, 40, 40, 40]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = larg

    # Legenda
    ws_leg = wb.create_sheet("Legenda")
    ws_leg["A1"] = "Cor"
    ws_leg["B1"] = "Significado"
    ws_leg["A2"].fill = verde
    ws_leg["A2"] = "Verde"
    ws_leg["B2"] = "PDF baixado automaticamente — arquivo na pasta pdfs/"
    ws_leg["A3"].fill = amarelo
    ws_leg["A3"] = "Amarelo"
    ws_leg["B3"] = "Sem PDF — acesso manual necessário (clique no link da coluna 'Acesso Manual')"
    ws_leg["A4"].fill = branco
    ws_leg["A4"] = "Branco"
    ws_leg["B4"] = "Sem link disponível"
    for col in ["A", "B"]:
        ws_leg.column_dimensions[col].width = 30

    wb.save(XLSX_FILE)
    log.info(f"Excel salvo: {XLSX_FILE}")


# ──────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("-q", "--query", required=True)
    parser.add_argument("-n", "--numero", type=int, default=111)
    parser.add_argument("--proxy")
    parser.add_argument("--sem-pdf", action="store_true")
    parser.add_argument("--resetar", action="store_true")
    args = parser.parse_args()

    if args.resetar:
        for f in [PROGRESSO, JSON_FILE, XLSX_FILE]:
            if f.exists():
                f.unlink()
        log.info("Resetado.")

    log.info("=" * 60)
    log.info(f"Query  : {args.query}")
    log.info(f"Filtro : até 2025")
    log.info(f"Meta   : {args.numero} artigos")
    log.info("=" * 60)

    sessao  = criar_sessao(proxy=args.proxy)
    artigos = buscar_tudo(sessao, args.query, args.numero)

    if not args.sem_pdf:
        artigos = baixar_todos_pdfs(sessao, artigos)

    salvar_xlsx(artigos)

    # JSON final
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(artigos, f, ensure_ascii=False, indent=2)

    pdfs_ok      = sum(1 for a in artigos if a.get("pdf_local"))
    manual       = sum(1 for a in artigos if a.get("acesso_manual"))
    sem_link     = len(artigos) - pdfs_ok - manual

    print(f"\n{'='*50}")
    print(f"✅ Concluído!")
    print(f"   Artigos coletados : {len(artigos)}")
    print(f"   🟢 PDFs baixados  : {pdfs_ok}")
    print(f"   🟡 Acesso manual  : {manual}  ← links no Excel, coluna 'Acesso Manual'")
    print(f"   ⚪ Sem link       : {sem_link}")
    print(f"   Pasta de saída    : {OUTPUT_DIR.resolve()}")
    print(f"   Planilha Excel    : {XLSX_FILE.resolve()}")
    print(f"{'='*50}")

if __name__ == "__main__":
    main()
